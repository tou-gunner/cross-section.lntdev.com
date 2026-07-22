#!/usr/bin/env python3
"""River cross-section survey processor.

Reads a raw survey CSV (No.,Northing,Easting,Elevation in UTM), splits it into
cross-section lines, reorders each section from the smaller-Easting bank to the
larger-Easting bank, renumbers points, and writes a static dataset folder
(JSON/GeoJSON for the web viewer + per-section CSV deliverables) plus a QC report.

An optional multibeam CSV (same column format) is processed as a second,
independent section stream: ids XMU-1..XMU-n, its own QC accounting, output to
multibeam_lines.json + sections/XMU-*.json + csv/XMU-*.csv in the same dataset.
No longitudinal-tail detection is applied to it — trailing spot points become
strays.

Pure stdlib (no numpy/pyproj). Python 3.10+.

Usage example:
    python3 tools/process.py \
        --input "Data/4_Cross section_Pakchom" \
        --csv "Cross Section_MSL.csv" \
        --check-csv "Cross Section_Hondau.csv" --check-offset 0.14 \
        --multibeam-csv "Section_multibeam-pakchom_MSL.csv" \
        --multibeam-check-csv "Section_multibeam-pakchom_HONDAU.csv" \
        --dataset-id pakchom --name "PakChom" \
        --out datasets/pakchom --index datasets/datasets.json
"""

import argparse
import csv
import datetime
import json
import math
import os
import sys
import zipfile

# ---------------------------------------------------------------------------
# Transverse Mercator (UTM) <-> WGS84, Krueger n-series to 6th order.
# Sub-centimeter accuracy even hundreds of km from the central meridian.
# ---------------------------------------------------------------------------

_A_WGS84 = 6378137.0
_F_WGS84 = 1.0 / 298.257223563
_K0 = 0.9996
_E0 = 500000.0


def _tm_constants():
    f = _F_WGS84
    n = f / (2.0 - f)
    n2, n3, n4, n5, n6 = n**2, n**3, n**4, n**5, n**6
    A = _A_WGS84 / (1 + n) * (1 + n2 / 4 + n4 / 64 + n6 / 256)
    alpha = [
        n / 2 - 2 * n2 / 3 + 5 * n3 / 16 + 41 * n4 / 180 - 127 * n5 / 288 + 7891 * n6 / 37800,
        13 * n2 / 48 - 3 * n3 / 5 + 557 * n4 / 1440 + 281 * n5 / 630 - 1983433 * n6 / 1935360,
        61 * n3 / 240 - 103 * n4 / 140 + 15061 * n5 / 26880 + 167603 * n6 / 181440,
        49561 * n4 / 161280 - 179 * n5 / 168 + 6601661 * n6 / 7257600,
        34729 * n5 / 80640 - 3418889 * n6 / 1995840,
        212378941 * n6 / 319334400,
    ]
    beta = [
        n / 2 - 2 * n2 / 3 + 37 * n3 / 96 - n4 / 360 - 81 * n5 / 512 + 96199 * n6 / 604800,
        n2 / 48 + n3 / 15 - 437 * n4 / 1440 + 46 * n5 / 105 - 1118711 * n6 / 3870720,
        17 * n3 / 480 - 37 * n4 / 840 - 209 * n5 / 4480 + 5569 * n6 / 90720,
        4397 * n4 / 161280 - 11 * n5 / 504 - 830251 * n6 / 7257600,
        4583 * n5 / 161280 - 108847 * n6 / 3991680,
        20648693 * n6 / 638668800,
    ]
    # conformal latitude chi -> geodetic latitude phi
    delta = [
        2 * n - 2 * n2 / 3 - 2 * n3 + 116 * n4 / 45 + 26 * n5 / 45 - 2854 * n6 / 675,
        7 * n2 / 3 - 8 * n3 / 5 - 227 * n4 / 45 + 2704 * n5 / 315 + 2323 * n6 / 945,
        56 * n3 / 15 - 136 * n4 / 35 - 1262 * n5 / 105 + 73814 * n6 / 2835,
        4279 * n4 / 630 - 332 * n5 / 35 - 399572 * n6 / 14175,
        4174 * n5 / 315 - 144838 * n6 / 6237,
        601676 * n6 / 22275,
    ]
    e = math.sqrt(f * (2 - f))
    return A, alpha, beta, delta, e


_TM_A, _TM_ALPHA, _TM_BETA, _TM_DELTA, _TM_E = _tm_constants()


def utm_to_wgs84(easting, northing, zone=48, northern=True):
    """Inverse UTM -> (lat, lng) in degrees."""
    lon0 = math.radians(zone * 6 - 183)
    n0 = 0.0 if northern else 10000000.0
    xi = (northing - n0) / (_K0 * _TM_A)
    eta = (easting - _E0) / (_K0 * _TM_A)
    xi_p, eta_p = xi, eta
    for j, b in enumerate(_TM_BETA, start=1):
        xi_p -= b * math.sin(2 * j * xi) * math.cosh(2 * j * eta)
        eta_p -= b * math.cos(2 * j * xi) * math.sinh(2 * j * eta)
    chi = math.atan2(math.sin(xi_p), math.hypot(math.sinh(eta_p), math.cos(xi_p)))
    lam = lon0 + math.atan2(math.sinh(eta_p), math.cos(xi_p))
    phi = chi
    for j, d in enumerate(_TM_DELTA, start=1):
        phi += d * math.sin(2 * j * chi)
    return math.degrees(phi), math.degrees(lam)


def wgs84_to_utm(lat, lng, zone=48, northern=True):
    """Forward UTM (used only for the round-trip self-test)."""
    lon0 = math.radians(zone * 6 - 183)
    phi = math.radians(lat)
    dlam = math.radians(lng) - lon0
    e = _TM_E
    t = math.sinh(math.asinh(math.tan(phi)) - e * math.atanh(e * math.sin(phi)))
    xi_p = math.atan2(t, math.cos(dlam))
    eta_p = math.asinh(math.sin(dlam) / math.hypot(t, math.cos(dlam)))
    xi, eta = xi_p, eta_p
    for j, a in enumerate(_TM_ALPHA, start=1):
        xi += a * math.sin(2 * j * xi_p) * math.cosh(2 * j * eta_p)
        eta += a * math.cos(2 * j * xi_p) * math.sinh(2 * j * eta_p)
    easting = _E0 + _K0 * _TM_A * eta
    northing = _K0 * _TM_A * xi + (0.0 if northern else 10000000.0)
    return easting, northing


def projection_selftest(points, zone, northern):
    """Round-trip a few real points; return worst error in meters."""
    worst = 0.0
    sample = [points[0], points[len(points) // 2], points[-1]]
    for p in sample:
        lat, lng = utm_to_wgs84(p["e"], p["n"], zone, northern)
        e2, n2 = wgs84_to_utm(lat, lng, zone, northern)
        worst = max(worst, math.hypot(e2 - p["e"], n2 - p["n"]))
    return worst


# ---------------------------------------------------------------------------
# Reading and cleaning
# ---------------------------------------------------------------------------

def fix_elevation(raw, fixes, no):
    """Parse an elevation cell; repair Excel percent-format exports."""
    s = raw.strip()
    if s.endswith("%"):
        val = float(s[:-1].replace(",", "")) / 100.0
        fixes.append({"no": no, "raw": s, "fixed": round(val, 3)})
        return val
    return float(s)


def _csv_rows(path):
    with open(path, newline="", encoding="utf-8-sig") as f:
        yield from csv.reader(f)


def _xlsx_rows(path):
    """First worksheet of an .xlsx as rows of strings (lazy openpyxl import)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit("reading .xlsx requires openpyxl (pip install openpyxl)"
                         " — or export the sheet to CSV")
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for row in wb.worksheets[0].iter_rows(values_only=True):
            yield ["" if c is None else str(c) for c in row[:4]]
    finally:
        wb.close()


def read_points(path, fixes, warnings, z_band=(100.0, 400.0)):
    rows = _xlsx_rows(path) if path.lower().endswith(".xlsx") else _csv_rows(path)
    pts = []
    header = next(rows)
    if len(header) < 4:
        raise SystemExit(f"unexpected header in {path}: {header}")
    for row in rows:
        if not row or not row[0].strip():
            continue
        no = int(float(row[0]))  # openpyxl may yield "13.0"
        z = fix_elevation(row[3], fixes, no)
        if not (z_band[0] <= z <= z_band[1]):
            warnings.append(f"point No.{no}: elevation {z} outside sanity band {z_band}")
        pts.append({"no": no, "n": float(row[1]), "e": float(row[2]), "z": z})
    return pts


def dist2d(a, b):
    return math.hypot(a["n"] - b["n"], a["e"] - b["e"])


# ---------------------------------------------------------------------------
# Structure detection
# ---------------------------------------------------------------------------

def split_on_gaps(points, max_gap):
    runs = [[points[0]]]
    for prev, cur in zip(points, points[1:]):
        if dist2d(prev, cur) > max_gap:
            runs.append([cur])
        else:
            runs[-1].append(cur)
    return runs


def detect_longitudinal_tail(points, max_gap, min_section_pts):
    """Everything after the last dense run (>= min_section_pts at <= max_gap
    spacing) is the longitudinal line surveyed after the sections.
    Returns (head_points, tail_points)."""
    runs = split_on_gaps(points, max_gap)
    last_big = None
    for i, r in enumerate(runs):
        if len(r) >= min_section_pts:
            last_big = i
    if last_big is None:
        return points, []
    head, tail = [], []
    for i, r in enumerate(runs):
        (head if i <= last_big else tail).extend(r)
    return head, tail


def classify_runs(runs, min_section_pts):
    sections, fragments = [], []
    for r in runs:
        (sections if len(r) >= min_section_pts else fragments).append(r)
    return sections, fragments


# ---------------------------------------------------------------------------
# Geometry: PCA line fit, fragment merging, ordering
# ---------------------------------------------------------------------------

def fit_line_pca(pts):
    """Principal axis through the points in the (E, N) plane.
    Returns (cE, cN, ux, uy) with (ux, uy) a unit direction vector."""
    ce = sum(p["e"] for p in pts) / len(pts)
    cn = sum(p["n"] for p in pts) / len(pts)
    see = sum((p["e"] - ce) ** 2 for p in pts)
    snn = sum((p["n"] - cn) ** 2 for p in pts)
    sen = sum((p["e"] - ce) * (p["n"] - cn) for p in pts)
    theta = 0.5 * math.atan2(2 * sen, see - snn)
    return ce, cn, math.cos(theta), math.sin(theta)


def station(p, line):
    ce, cn, ux, uy = line
    return (p["e"] - ce) * ux + (p["n"] - cn) * uy


def perp_dist(p, line):
    ce, cn, ux, uy = line
    return abs(-(p["e"] - ce) * uy + (p["n"] - cn) * ux)


def join_collinear_sections(sections, bearing_tol, perp_tol, max_gap, qc):
    """Union section runs that are collinear pieces of one physical
    cross-section (a main-channel line plus overbank 'wing' lines surveyed
    as separate runs, possibly in a distant block of the file).

    Pair rule: PCA bearing difference <= bearing_tol (deg, mod 180), mutual
    centroid-to-line perpendicular distance <= perp_tol (m), and minimum
    end-to-end gap <= max_gap (m); grouping is transitive. Distinct adjacent
    cross-sections are parallel but laterally offset, so they never pass the
    perp test. Near-miss pairs are reported as warnings for human review."""
    n = len(sections)
    if n < 2:
        return sections

    geom = []
    for sec in sections:
        line = sec["line"]
        st = [station(p, line) for p in sec["pts"]]
        geom.append({
            "brg": math.degrees(math.atan2(line[3], line[2])) % 180.0,
            "c": {"e": line[0], "n": line[1]},
            "ends": [sec["pts"][st.index(min(st))], sec["pts"][st.index(max(st))]],
        })

    def bearing_diff(a, b):
        d = abs(a - b) % 180.0
        return min(d, 180.0 - d)

    parent = list(range(n))

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    nearmiss = []
    for i in range(n):
        gi = geom[i]
        for j in range(i + 1, n):
            gj = geom[j]
            bd = bearing_diff(gi["brg"], gj["brg"])
            if bd > bearing_tol:
                continue
            pp = max(perp_dist(gj["c"], sections[i]["line"]),
                     perp_dist(gi["c"], sections[j]["line"]))
            gap = min(dist2d(p, q) for p in gi["ends"] for q in gj["ends"])
            if pp <= perp_tol and gap <= max_gap:
                parent[find(i)] = find(j)
            elif pp <= perp_tol and gap <= 2 * max_gap:
                nearmiss.append(("gap", i, j, bd, pp, gap))
            elif pp <= 2 * perp_tol and gap <= max_gap:
                nearmiss.append(("perp", i, j, bd, pp, gap))

    groups = {}
    for i in range(n):
        groups.setdefault(find(i), []).append(i)

    result, joined = [], []
    for members in groups.values():
        if len(members) == 1:
            result.append(sections[members[0]])  # untouched original dict
            continue
        pieces = sorted((sections[i] for i in members), key=lambda s: s["min_no"])
        new_pts = [p for s in pieces for p in s["pts"]]
        line = fit_line_pca(new_pts)
        st = [station(p, line) for p in new_pts]
        intervals = sorted(
            (min(ps), max(ps)) for ps in
            ([station(p, line) for p in s["pts"]] for s in pieces))
        joined.append({
            "piece_start_nos": [s["min_no"] for s in pieces],
            "piece_points": [len(s["pts"]) for s in pieces],
            "bearing_spread_deg": round(max(
                bearing_diff(geom[i]["brg"], geom[j]["brg"])
                for i in members for j in members), 2),
            "max_perp_m": round(max(
                perp_dist(geom[i]["c"], line) for i in members), 2),
            "gaps_m": [round(max(0.0, b[0] - a[1]), 1)
                       for a, b in zip(intervals, intervals[1:])],
        })
        result.append({"pts": new_pts, "line": line,
                       "smin": min(st), "smax": max(st),
                       "min_no": min(s["min_no"] for s in pieces), "merged": []})

    if joined:
        joined.sort(key=lambda g: g["piece_start_nos"][0])
        qc["joined_sections"] = {
            "params": {"bearing_tol_deg": bearing_tol, "perp_tol_m": perp_tol,
                       "max_gap_m": max_gap},
            "groups": joined,
        }

    seen = set()
    for kind, i, j, bd, pp, gap in nearmiss:
        ri, rj = find(i), find(j)
        if ri == rj:
            continue  # transitively joined anyway — not a real miss
        key = (kind, min(ri, rj), max(ri, rj))
        if key in seen:
            continue
        seen.add(key)
        a, b = sections[i]["min_no"], sections[j]["min_no"]
        if kind == "gap":
            qc["warnings"].append(
                f"join near-miss (gap): sections starting No.{a} and No.{b} "
                f"collinear (bearing diff {bd:.2f} deg, perp {pp:.1f} m) but end "
                f"gap {gap:.0f} m > --join-max-gap {max_gap:.0f} m — verify "
                "whether one cross-section")
        else:
            qc["warnings"].append(
                f"join near-miss (perp): sections starting No.{a} and No.{b} "
                f"nearly collinear (bearing diff {bd:.2f} deg, end gap {gap:.0f} m) "
                f"but perp {pp:.1f} m > --join-perp-tol {perp_tol:.0f} m — verify "
                "whether one cross-section")

    result.sort(key=lambda s: s["min_no"])
    return result


def merge_fragments(sections, fragments, perp_tol, ext_tol, qc):
    """Attach each small run to a collinear neighboring section, else stray.
    `sections` is a list of dicts {pts, line, smin, smax, min_no}."""
    strays = []
    for frag in fragments:
        frag_min_no = min(p["no"] for p in frag)
        # candidates ordered by closeness in survey (file) order
        cands = sorted(sections, key=lambda s: abs(s["min_no"] - frag_min_no))
        merged = False
        best_reason, best_sec, best_perp = "no_candidate", None, None
        for sec in cands:
            perps = [perp_dist(p, sec["line"]) for p in frag]
            stations = [station(p, sec["line"]) for p in frag]
            pmax = max(perps)
            if best_perp is None or pmax < best_perp:
                best_perp, best_sec = pmax, sec
            if pmax >= perp_tol:
                best_reason = "perp_too_far"
                continue
            if not (sec["smin"] - ext_tol <= min(stations) and max(stations) <= sec["smax"] + ext_tol):
                best_reason = "station_out_of_range"
                continue
            sec["pts"].extend(frag)
            sec["merged"].append({"nos": [p["no"] for p in frag], "max_perp_m": round(pmax, 2)})
            qc["merged_fragments"].append({
                "fragment_nos": [p["no"] for p in frag],
                "into_min_no": sec["min_no"],
                "max_perp_m": round(pmax, 2),
            })
            merged = True
            break
        if not merged:
            for p in frag:
                strays.append({
                    **p,
                    "reason": best_reason,
                    "nearest_section_min_no": best_sec["min_no"] if best_sec else None,
                    "perp_dist_m": round(best_perp, 2) if best_perp is not None else None,
                })
    return strays


def sort_along_line(sec):
    """Re-fit the line after merging and sort points along it
    (direction not yet meaningful — fixed later by orient_sections)."""
    line = fit_line_pca(sec["pts"])
    sec["pts"].sort(key=lambda p: station(p, line))
    sec["line"] = line


def orient_sections(sections, gap_thresh, qc):
    """Orient every section to start on the LEFT bank looking downstream.

    Sections must already be in along-river order. The local downstream
    direction at section i is taken from the neighboring sections' centroids;
    rotating it 90° clockwise gives a vector toward the right bank, so the
    section's start->end direction must agree with it (start = left bank).
    """
    max_link = 3000.0  # file-neighbor links longer than this are survey jumps, not river adjacency
    cents = [(s["line"][0], s["line"][1]) for s in sections]

    def link_ok(i, j):
        return 0 <= j < len(cents) and math.hypot(
            cents[j][0] - cents[i][0], cents[j][1] - cents[i][1]) < max_link

    def nearest_segment_flow(i):
        """Flow for a section with no trusted neighbors: direction of the
        nearest segment of the ordered-centroid polyline (skipping itself)."""
        best, best_d = None, float("inf")
        px, py = cents[i]
        for j in range(len(cents) - 1):
            if j == i or j + 1 == i:
                continue
            ax, ay = cents[j]
            bx, by = cents[j + 1]
            vx, vy = bx - ax, by - ay
            L2 = vx * vx + vy * vy
            if L2 == 0:
                continue
            u = max(0.0, min(1.0, ((px - ax) * vx + (py - ay) * vy) / L2))
            d = math.hypot(px - (ax + u * vx), py - (ay + u * vy))
            if d < best_d:
                best_d, best = d, (j, (vx, vy))
        return best, best_d

    flipped = []
    for i, sec in enumerate(sections):
        prev_ok = link_ok(i, i - 1)
        next_ok = link_ok(i, i + 1)
        if prev_ok or next_ok:
            a = cents[i - 1] if prev_ok else cents[i]
            b = cents[i + 1] if next_ok else cents[i]
            flow = (b[0] - a[0], b[1] - a[1])      # (E, N) downstream
        else:
            seg, d = nearest_segment_flow(i)
            if seg is None:
                flow = (0.0, 0.0)
            else:
                flow = seg[1]
                qc["warnings"].append(
                    f"section starting No.{sec['min_no']}: isolated (no neighbor within "
                    f"{max_link:.0f} m); flow taken from chain segment after section "
                    f"index {seg[0] + 1} ({d:,.0f} m away) — verify bank orientation")
        rightvec = (flow[1], -flow[0])             # toward right bank
        pts = sec["pts"]
        u = (pts[-1]["e"] - pts[0]["e"], pts[-1]["n"] - pts[0]["n"])
        dot = u[0] * rightvec[0] + u[1] * rightvec[1]
        if dot == 0:
            qc["warnings"].append(
                f"section starting No.{sec['min_no']}: flow-orientation ambiguous, "
                "falling back to smaller-Easting start")
            if pts[0]["e"] > pts[-1]["e"]:
                pts.reverse()
                flipped.append(sec["min_no"])
        elif dot < 0:
            pts.reverse()
            flipped.append(sec["min_no"])
        offsets, gaps = [0.0], []
        for j, (prev, cur) in enumerate(zip(pts, pts[1:]), start=1):
            d = dist2d(prev, cur)
            offsets.append(offsets[-1] + d)
            if d > gap_thresh:
                gaps.append({"after_no": j, "length_m": round(d, 1)})
        sec["ordered"] = pts
        sec["offsets"] = offsets
        sec["gaps"] = gaps
    qc["orientation"] = {"rule": "left_bank_downstream", "flipped": len(flipped),
                         "flipped_section_start_nos": flipped}


def build_section_stream(points, args, qc):
    """split -> classify -> (join) -> merge fragments -> sort along line.

    The shared machinery for one survey stream (main or multibeam). `qc` must
    have 'warnings' and 'merged_fragments' lists pre-created; join and merge
    results are namespaced into it. Returns (sections, strays)."""
    runs = split_on_gaps(points, args.gap)
    sec_runs, fragments = classify_runs(runs, args.min_section_pts)
    sections = []
    for r in sec_runs:
        line = fit_line_pca(r)
        st = [station(p, line) for p in r]
        sections.append({"pts": r, "line": line, "smin": min(st), "smax": max(st),
                         "min_no": min(p["no"] for p in r), "merged": []})

    if args.join_collinear:
        sections = join_collinear_sections(
            sections, args.join_bearing_tol, args.join_perp_tol,
            args.join_max_gap, qc)

    strays = merge_fragments(sections, fragments, args.merge_perp, args.merge_ext, qc)

    for sec in sections:
        sort_along_line(sec)
    sections.sort(key=lambda s: min(p["no"] for p in s["pts"]))
    return sections, strays


def order_along_chain(sections, ref_sections, qc):
    """Put `sections` (in place) into along-river order so ids follow the
    river and orient_sections sees true neighbors.

    Two steps: (1) repair out-of-sequence surveying by moving single sections
    to the position that minimizes total centroid-chain length (file order is
    kept wherever it is already shortest — the sections are dense and evenly
    spaced, so the shortest open path through the centroids IS the along-river
    order); (2) reverse the whole list if it runs opposite to the reference
    (main-survey) chain, so numbering and the left-bank-downstream convention
    match across streams. Must run before orient_sections."""
    n = len(sections)
    qc["ordering"] = {"moved": 0, "reversed": False}
    if n < 2:
        return
    cents = [(s["line"][0], s["line"][1]) for s in sections]

    def total(order):
        return sum(math.hypot(cents[b][0] - cents[a][0], cents[b][1] - cents[a][1])
                   for a, b in zip(order, order[1:]))

    order = list(range(n))
    best = total(order)
    moved = 0
    improved = True
    while improved and moved < n:  # n moves is far beyond any real repair
        improved = False
        for i in range(n):
            rest = order[:i] + order[i + 1:]
            for j in range(n):
                if j == i:
                    continue
                cand = rest[:j] + [order[i]] + rest[j:]
                t = total(cand)
                if t < best - 1e-6:
                    order, best, improved = cand, t, True
                    moved += 1
                    break
            if improved:
                break
    if moved:
        sections[:] = [sections[i] for i in order]
        qc["ordering"]["moved"] = moved
        qc["warnings"].append(
            f"file order does not follow the river — {moved} section move(s) "
            "applied to restore along-river order")

    if len(ref_sections) >= 2:
        ve = sections[-1]["line"][0] - sections[0]["line"][0]
        vn = sections[-1]["line"][1] - sections[0]["line"][1]
        we = ref_sections[-1]["line"][0] - ref_sections[0]["line"][0]
        wn = ref_sections[-1]["line"][1] - ref_sections[0]["line"][1]
        if ve * we + vn * wn < 0:
            sections.reverse()
            qc["ordering"]["reversed"] = True
            qc["warnings"].append(
                "chain runs opposite to the main survey — section list "
                "reversed so numbering matches the along-river convention")


def datum_check(points, input_dir, check_csv, offset, z_band):
    """Per-point elevation comparison against a second-datum export of the
    same survey (keyed by point No.; coordinates are not compared)."""
    cf, cw = [], []
    cpath = os.path.join(input_dir, check_csv)
    cpoints = {p["no"]: p for p in read_points(cpath, cf, cw, z_band)}
    max_dev, over, rows = 0.0, 0, 0
    for p in points:
        q = cpoints.get(p["no"])
        if q is None:
            continue
        rows += 1
        dev = abs((p["z"] - q["z"]) - offset)
        max_dev = max(max_dev, dev)
        if dev > 0.01:
            over += 1
    return {"file": check_csv, "expected_offset": offset,
            "rows": rows, "max_dev_m": round(max_dev, 4),
            "rows_over_1cm": over}


# ---------------------------------------------------------------------------
# Longitudinal line
# ---------------------------------------------------------------------------

def build_longitudinal(tail, seg_gap=2000.0):
    """Split the tail into display segments at very large jumps; chainage is
    cumulative along the whole tail (including jumps)."""
    if not tail:
        return {"segments": [], "chainage": [], "points": 0, "length_m": 0.0}
    segments = [[tail[0]]]
    chain = [0.0]
    for prev, cur in zip(tail, tail[1:]):
        d = dist2d(prev, cur)
        chain.append(chain[-1] + d)
        if d > seg_gap:
            segments.append([cur])
        else:
            segments[-1].append(cur)
    return {"segments": segments, "chainage": chain, "points": len(tail),
            "length_m": round(chain[-1], 1)}


# ---------------------------------------------------------------------------
# Writers
# ---------------------------------------------------------------------------

def jdump(obj, path, indent=None):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=indent,
                  separators=None if indent else (",", ":"))


# ZIP epoch (1980-01-01) — the minimum a ZIP entry can store. Using a fixed
# timestamp keeps the archive byte-stable across reruns when the section CSVs
# are unchanged, instead of churning on every build.
_ZIP_EPOCH = (1980, 1, 1, 0, 0, 0)


def write_csv_zip(out_dir, dataset_id, section_ids):
    """Bundle every per-section CSV into one downloadable archive.

    Entries are stored under a `<id>_csv/` folder so they extract into a tidy
    directory rather than scattering 100+ loose files into the user's Downloads.
    """
    zip_path = os.path.join(out_dir, "all_csv.zip")
    folder = f"{dataset_id}_csv"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED, compresslevel=9) as zf:
        for sid in section_ids:
            with open(os.path.join(out_dir, "csv", f"{sid}.csv"), "rb") as fh:
                data = fh.read()
            info = zipfile.ZipInfo(f"{folder}/{sid}.csv", date_time=_ZIP_EPOCH)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = 0o644 << 16  # regular file, rw-r--r--
            zf.writestr(info, data)
    return {"file": "all_csv.zip", "files": len(section_ids),
            "bytes": os.path.getsize(zip_path)}


def write_section_set(out_dir, sections, prefix, zone, northern, datum,
                      lat_all, lng_all):
    """Write per-section JSON + CSV for one section stream (ids <prefix>-<idx>).
    Returns (manifest_sections, line_features, section_ids)."""
    manifest_sections = []
    line_features = []
    section_ids = []

    for idx, sec in enumerate(sections, start=1):
        sid = f"{prefix}-{idx}"
        section_ids.append(sid)
        pts, offs = sec["ordered"], sec["offsets"]
        lats, lngs = [], []
        for p in pts:
            lat, lng = utm_to_wgs84(p["e"], p["n"], zone, northern)
            lats.append(round(lat, 6))
            lngs.append(round(lng, 6))
        lat_all.extend(lats)
        lng_all.extend(lngs)
        zs = [p["z"] for p in pts]
        sec_json = {
            "id": sid,
            "datum": datum,
            "no": list(range(1, len(pts) + 1)),
            "offset": [round(o, 2) for o in offs],
            "z": [round(z, 3) for z in zs],
            "n": [round(p["n"], 3) for p in pts],
            "e": [round(p["e"], 3) for p in pts],
            "lat": lats,
            "lng": lngs,
            "gaps": sec["gaps"],
            "orig_no": [p["no"] for p in pts],
        }
        jdump(sec_json, os.path.join(out_dir, "sections", f"{sid}.json"))

        csv_path = os.path.join(out_dir, "csv", f"{sid}.csv")
        os.makedirs(os.path.dirname(csv_path), exist_ok=True)
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f, lineterminator="\r\n")
            w.writerow(["No", "Offset", "Northing", "Easting", "Elevation"])
            for i, p in enumerate(pts):
                w.writerow([i + 1, f"{offs[i]:.2f}", f"{p['n']:.3f}",
                            f"{p['e']:.3f}", f"{p['z']:.3f}"])

        mid = len(pts) // 2
        manifest_sections.append({
            "id": sid,
            "points": len(pts),
            "length_m": round(offs[-1], 1),
            "zmin": round(min(zs), 3),
            "zmax": round(max(zs), 3),
            "mid": [lats[mid], lngs[mid]],
            "merged_fragments": len(sec["merged"]),
            "file": f"sections/{sid}.json",
            "csv": f"csv/{sid}.csv",
        })
        line_features.append({
            "type": "Feature",
            "properties": {"id": sid, "points": len(pts), "length_m": round(offs[-1], 1)},
            "geometry": {"type": "LineString",
                         "coordinates": [[lngs[i], lats[i]] for i in range(len(pts))]},
        })

    return manifest_sections, line_features, section_ids


def write_outputs(out_dir, dataset_id, name, sections, mb_sections, longi,
                  strays, mb_strays, qc, args):
    zone, northern = args.utm_zone, args.hemisphere.upper() == "N"
    lat_all, lng_all = [], []

    manifest_sections, line_features, section_ids = write_section_set(
        out_dir, sections, "X", zone, northern, args.datum, lat_all, lng_all)
    jdump({"type": "FeatureCollection", "features": line_features},
          os.path.join(out_dir, "lines.json"))

    # multibeam stream: always write the file (empty when absent) so the
    # viewer's fetch never 404s on a fresh rebuild
    mb_manifest_sections, mb_line_features, mb_ids = write_section_set(
        out_dir, mb_sections, "XMU", zone, northern, args.datum, lat_all, lng_all)
    jdump({"type": "FeatureCollection", "features": mb_line_features},
          os.path.join(out_dir, "multibeam_lines.json"))

    # longitudinal line
    longi_features = []
    ci = 0
    for seg in longi["segments"]:
        coords, chain, zs = [], [], []
        for p in seg:
            lat, lng = utm_to_wgs84(p["e"], p["n"], zone, northern)
            lat, lng = round(lat, 6), round(lng, 6)
            coords.append([lng, lat])
            lat_all.append(lat)
            lng_all.append(lng)
            chain.append(round(longi["chainage"][ci], 1))
            zs.append(round(p["z"], 3))
            ci += 1
        longi_features.append({
            "type": "Feature",
            "properties": {"chainage": chain, "z": zs},
            "geometry": {"type": "LineString", "coordinates": coords},
        })
    jdump({"type": "FeatureCollection", "features": longi_features},
          os.path.join(out_dir, "longitudinal.json"))

    # strays (multibeam-origin points carry a "source" property)
    stray_features = []
    for source, batch in ((None, strays), ("multibeam", mb_strays)):
        for s in batch:
            lat, lng = utm_to_wgs84(s["e"], s["n"], zone, northern)
            lat, lng = round(lat, 6), round(lng, 6)
            lat_all.append(lat)
            lng_all.append(lng)
            props = {"no": s["no"], "n": round(s["n"], 3), "e": round(s["e"], 3),
                     "z": round(s["z"], 3), "reason": s["reason"],
                     "nearest_section_min_no": s["nearest_section_min_no"],
                     "perp_dist_m": s["perp_dist_m"]}
            if source:
                props["source"] = source
            stray_features.append({
                "type": "Feature",
                "properties": props,
                "geometry": {"type": "Point", "coordinates": [lng, lat]},
            })
    jdump({"type": "FeatureCollection", "features": stray_features},
          os.path.join(out_dir, "strays.json"))

    csv_zip = write_csv_zip(out_dir, dataset_id, section_ids + mb_ids)

    section_points = sum(len(s["ordered"]) for s in sections)
    mb_section_points = sum(len(s["ordered"]) for s in mb_sections)
    mb_qc = qc.get("multibeam")
    manifest = {
        "id": dataset_id,
        "name": name,
        "crs": f"EPSG:{'326' if northern else '327'}{zone}",
        "datum": args.datum,
        "orientation": "left_bank_downstream",
        "bounds": [[round(min(lat_all), 6), round(min(lng_all), 6)],
                   [round(max(lat_all), 6), round(max(lng_all), 6)]],
        "counts": {"sections": len(sections), "section_points": section_points,
                   "longitudinal_points": longi["points"],
                   "stray_points": len(strays) + len(mb_strays),
                   "multibeam_sections": len(mb_sections),
                   "multibeam_section_points": mb_section_points,
                   "multibeam_stray_points": len(mb_strays),
                   "total": qc["input_rows"] + (mb_qc["input_rows"] if mb_qc else 0)},
        "sections": manifest_sections,
        "multibeam_sections": mb_manifest_sections,
        "multibeam_lines": "multibeam_lines.json",
        "longitudinal": {"file": "longitudinal.json", "points": longi["points"],
                         "length_m": longi["length_m"]},
        "strays": {"file": "strays.json", "points": len(strays) + len(mb_strays)},
        "csv_zip": csv_zip,
        "qc": "qc_report.json",
    }
    jdump(manifest, os.path.join(out_dir, "manifest.json"), indent=2)
    return manifest


def write_qc(out_dir, qc):
    jdump(qc, os.path.join(out_dir, "qc_report.json"), indent=2)
    lines = [
        "QC REPORT — cross-section pipeline",
        f"generated: {qc['generated']}",
        f"input file: {qc['input_file']}",
        f"input rows: {qc['input_rows']}",
        "",
        f"elevation fixes (percent-format cells): {len(qc['elevation_fixes'])}",
        *(f"  No.{f['no']}: {f['raw']} -> {f['fixed']}" for f in qc["elevation_fixes"]),
        "",
        f"longitudinal tail: No.{qc['tail']['start_no']}–{qc['tail']['end_no']}"
        f" ({qc['tail']['points']} pts, {qc['tail']['length_m']} m)",
        f"sections: {qc['sections']}",
        f"orientation: start on left bank looking downstream"
        f" ({qc['orientation']['flipped']} sections flipped)",
        f"merged fragments: {len(qc['merged_fragments'])}",
        *(f"  Nos {m['fragment_nos']} -> section starting No.{m['into_min_no']}"
          f" (max perp {m['max_perp_m']} m)" for m in qc["merged_fragments"]),
    ]
    js = qc.get("joined_sections")
    if js:
        p = js["params"]
        lines += ["", f"joined collinear sections: {len(js['groups'])} groups "
                  f"(bearing tol {p['bearing_tol_deg']} deg, perp tol "
                  f"{p['perp_tol_m']} m, max gap {p['max_gap_m']} m)"]
        lines += [f"  start Nos {g['piece_start_nos']} (pts {g['piece_points']})"
                  f" -> one section (bearing spread {g['bearing_spread_deg']} deg,"
                  f" max perp {g['max_perp_m']} m, gaps {g['gaps_m']} m)"
                  for g in js["groups"]]
    lines += [
        "",
        f"stray points: {qc['stray_points']}",
        f"projection round-trip worst error: {qc['projection_selftest_m']} m",
    ]
    if qc.get("check"):
        c = qc["check"]
        lines += ["", f"datum cross-check vs {c['file']} (expected offset {c['expected_offset']} m):",
                  f"  rows compared: {c['rows']}, max deviation: {c['max_dev_m']} m,"
                  f" rows > 0.01 m: {c['rows_over_1cm']}"]
    mb = qc.get("multibeam")
    if mb:
        lines += [
            "",
            f"multibeam input: {mb['input_file']} ({mb['input_rows']} rows)",
            f"multibeam sections: {mb['sections']}"
            + (f" — {mb['ordering']['moved']} out-of-sequence move(s)"
               if mb.get("ordering", {}).get("moved") else "")
            + (" — list reversed to match main chain direction"
               if mb.get("ordering", {}).get("reversed") else ""),
            f"multibeam orientation: {mb['orientation']['flipped']} sections flipped",
            f"multibeam merged fragments: {len(mb['merged_fragments'])}",
            *(f"  Nos {m['fragment_nos']} -> section starting No.{m['into_min_no']}"
              f" (max perp {m['max_perp_m']} m)" for m in mb["merged_fragments"]),
            f"multibeam stray points: {mb['stray_points']}",
            f"multibeam projection round-trip worst error: {mb['projection_selftest_m']} m",
        ]
        mjs = mb.get("joined_sections")
        if mjs:
            p = mjs["params"]
            lines += [f"multibeam joined collinear sections: {len(mjs['groups'])} groups "
                      f"(bearing tol {p['bearing_tol_deg']} deg, perp tol "
                      f"{p['perp_tol_m']} m, max gap {p['max_gap_m']} m)"]
            lines += [f"  start Nos {g['piece_start_nos']} (pts {g['piece_points']})"
                      f" -> one section (bearing spread {g['bearing_spread_deg']} deg,"
                      f" max perp {g['max_perp_m']} m, gaps {g['gaps_m']} m)"
                      for g in mjs["groups"]]
        if mb.get("check"):
            c = mb["check"]
            lines += [f"multibeam datum cross-check vs {c['file']}"
                      f" (expected offset {c['expected_offset']} m):",
                      f"  rows compared: {c['rows']}, max deviation: {c['max_dev_m']} m,"
                      f" rows > 0.01 m: {c['rows_over_1cm']}"]
        if mb["warnings"]:
            lines += ["multibeam warnings:"] + [f"  {w}" for w in mb["warnings"]]
    if qc["warnings"]:
        lines += ["", "warnings:"] + [f"  {w}" for w in qc["warnings"]]
    lines += ["", f"INVARIANT section+longitudinal+stray == input: "
              f"{qc['invariant']['sum']} == {qc['input_rows']} -> {qc['invariant']['ok']}"]
    if mb:
        lines += [f"INVARIANT multibeam section+stray == input: "
                  f"{mb['invariant']['sum']} == {mb['input_rows']} -> {mb['invariant']['ok']}"]
    with open(os.path.join(out_dir, "qc_report.txt"), "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")


def update_datasets_index(index_path, dataset_id, name, out_dir, crs, datum):
    idx = {"version": 1, "datasets": []}
    if os.path.exists(index_path):
        with open(index_path, encoding="utf-8") as f:
            idx = json.load(f)
    web_root = os.path.dirname(os.path.abspath(index_path))
    rel = os.path.relpath(os.path.abspath(out_dir), os.path.dirname(web_root))
    entry = {"id": dataset_id, "name": name, "path": rel.replace(os.sep, "/"),
             "crs": crs, "datum": datum,
             # full timestamp: the viewer uses this as a cache-busting key,
             # so same-day rebuilds must produce a fresh value
             "updated": datetime.datetime.now().isoformat(timespec="seconds")}
    idx["datasets"] = [d for d in idx["datasets"] if d["id"] != dataset_id] + [entry]
    jdump(idx, index_path, indent=2)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--input", required=True, help="folder containing the raw CSV")
    ap.add_argument("--csv", required=True, help="CSV filename inside --input")
    ap.add_argument("--check-csv", help="second-datum CSV for cross-checking")
    ap.add_argument("--check-offset", type=float, default=0.0,
                    help="expected (main - check) elevation offset in meters")
    ap.add_argument("--multibeam-csv",
                    help="multibeam survey CSV filename inside --input "
                         "(processed as a second section stream, ids XMU-n)")
    ap.add_argument("--multibeam-check-csv",
                    help="second-datum multibeam CSV (uses --check-offset)")
    ap.add_argument("--dataset-id", required=True)
    ap.add_argument("--name", required=True, help="display name (may be Lao)")
    ap.add_argument("--out", required=True, help="dataset output folder")
    ap.add_argument("--index", help="datasets.json to upsert")
    ap.add_argument("--utm-zone", type=int, default=48)
    ap.add_argument("--hemisphere", default="N", choices=["N", "S"])
    ap.add_argument("--datum", default="MSL")
    ap.add_argument("--z-min", type=float, default=100.0,
                    help="elevation sanity-band floor in meters (default 100)")
    ap.add_argument("--z-max", type=float, default=400.0,
                    help="elevation sanity-band ceiling in meters (default 400)")
    ap.add_argument("--gap", type=float, default=50.0, help="section split gap (m)")
    ap.add_argument("--min-section-pts", type=int, default=10)
    ap.add_argument("--merge-perp", type=float, default=15.0)
    ap.add_argument("--merge-ext", type=float, default=200.0)
    ap.add_argument("--join-collinear", action=argparse.BooleanOptionalAction,
                    default=True,
                    help="join collinear section pieces (channel + overbank "
                         "wings) into one cross-section")
    ap.add_argument("--join-bearing-tol", type=float, default=3.0,
                    help="join: max PCA bearing difference between pieces (deg)")
    ap.add_argument("--join-perp-tol", type=float, default=25.0,
                    help="join: max mutual centroid-to-line perpendicular "
                         "distance (m); near-misses up to 2x are QC-warned — "
                         "raise after visual inspection if flagged")
    ap.add_argument("--join-max-gap", type=float, default=3000.0,
                    help="join: max end-to-end gap between collinear pieces (m)")
    args = ap.parse_args()

    qc = {"generated": datetime.datetime.now().isoformat(timespec="seconds"),
          "input_file": args.csv, "elevation_fixes": [], "merged_fragments": [],
          "warnings": []}

    path = os.path.join(args.input, args.csv)
    z_band = (args.z_min, args.z_max)
    points = read_points(path, qc["elevation_fixes"], qc["warnings"], z_band)
    qc["input_rows"] = len(points)
    nos = [p["no"] for p in points]
    if len(set(nos)) != len(nos):
        raise SystemExit("duplicate point numbers in input")

    northern = args.hemisphere.upper() == "N"
    qc["projection_selftest_m"] = round(
        projection_selftest(points, args.utm_zone, northern), 6)
    if qc["projection_selftest_m"] > 0.01:
        raise SystemExit(f"projection self-test failed: {qc['projection_selftest_m']} m")

    head, tail = detect_longitudinal_tail(points, args.gap, args.min_section_pts)
    longi = build_longitudinal(tail)
    qc["tail"] = {"start_no": tail[0]["no"] if tail else None,
                  "end_no": tail[-1]["no"] if tail else None,
                  "points": longi["points"], "length_m": longi["length_m"]}

    sections, strays = build_section_stream(head, args, qc)
    orient_sections(sections, args.gap, qc)
    qc["sections"] = len(sections)
    qc["stray_points"] = len(strays)

    # datum cross-check
    if args.check_csv:
        qc["check"] = datum_check(points, args.input, args.check_csv,
                                  args.check_offset, z_band)

    section_points = sum(len(s["ordered"]) for s in sections)
    total = section_points + longi["points"] + len(strays)
    qc["invariant"] = {"sum": total, "ok": total == qc["input_rows"]}
    if not qc["invariant"]["ok"]:
        raise SystemExit(f"POINT ACCOUNTING FAILED: {total} != {qc['input_rows']}")

    # multibeam stream (optional): same machinery, separate accounting.
    # No longitudinal-tail detection — trailing spot points become
    # fragments -> strays.
    mb_sections, mb_strays, mb_qc = [], [], None
    if args.multibeam_csv:
        mb_qc = {"input_file": args.multibeam_csv, "elevation_fixes": [],
                 "merged_fragments": [], "warnings": []}
        mb_points = read_points(os.path.join(args.input, args.multibeam_csv),
                                mb_qc["elevation_fixes"], mb_qc["warnings"], z_band)
        mb_qc["input_rows"] = len(mb_points)
        if not mb_points:
            raise SystemExit(f"multibeam input has no data rows: {args.multibeam_csv}")
        mb_nos = [p["no"] for p in mb_points]
        if len(set(mb_nos)) != len(mb_nos):
            raise SystemExit("duplicate point numbers in multibeam input")
        mb_qc["projection_selftest_m"] = round(
            projection_selftest(mb_points, args.utm_zone, northern), 6)
        if mb_qc["projection_selftest_m"] > 0.01:
            raise SystemExit("multibeam projection self-test failed: "
                             f"{mb_qc['projection_selftest_m']} m")
        mb_sections, mb_strays = build_section_stream(mb_points, args, mb_qc)
        order_along_chain(mb_sections, sections, mb_qc)
        orient_sections(mb_sections, args.gap, mb_qc)
        mb_qc["sections"] = len(mb_sections)
        mb_qc["stray_points"] = len(mb_strays)
        if args.multibeam_check_csv:
            mb_qc["check"] = datum_check(mb_points, args.input,
                                         args.multibeam_check_csv,
                                         args.check_offset, z_band)
        mb_total = sum(len(s["ordered"]) for s in mb_sections) + len(mb_strays)
        mb_qc["invariant"] = {"sum": mb_total, "ok": mb_total == mb_qc["input_rows"]}
        if not mb_qc["invariant"]["ok"]:
            raise SystemExit(
                f"MULTIBEAM POINT ACCOUNTING FAILED: {mb_total} != {mb_qc['input_rows']}")
        qc["multibeam"] = mb_qc

    manifest = write_outputs(args.out, args.dataset_id, args.name,
                             sections, mb_sections, longi, strays, mb_strays,
                             qc, args)
    write_qc(args.out, qc)
    if args.index:
        update_datasets_index(args.index, args.dataset_id, args.name,
                              args.out, manifest["crs"], args.datum)

    joined_n = len(qc.get("joined_sections", {}).get("groups", []))
    print(f"OK: {qc['sections']} sections, {section_points} section points, "
          f"{longi['points']} longitudinal, {len(strays)} strays, "
          f"{len(qc['elevation_fixes'])} elevation fixes"
          + (f", {joined_n} collinear joins" if joined_n else ""))
    if mb_qc:
        print(f"OK multibeam: {mb_qc['sections']} sections, "
              f"{sum(len(s['ordered']) for s in mb_sections)} section points, "
              f"{len(mb_strays)} strays"
              + (", reordered" if mb_qc.get("ordering", {}).get("moved") else "")
              + (", reversed" if mb_qc.get("ordering", {}).get("reversed") else ""))
    print(f"QC report: {os.path.join(args.out, 'qc_report.txt')}")


if __name__ == "__main__":
    main()
